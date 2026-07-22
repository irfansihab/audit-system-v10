"""Tool fill_lke — isi Lembar Kerja Evaluasi (LKE) Excel untuk skill evaluasi
(SAKIP/SPIP) TANPA mengubah rumus. Hanya cell input yang ditulis; cell formula
& sheet agregator DITOLAK.

Sumber LKE:
  - SPIP: template ber-rumus yang dibawa skill (knowledge/skills/evaluasi-spip/
    references/templates/lke-spip-kementerian.xlsx + cell-map).
  - SAKIP / lainnya: LKE .xlsx yang DIUPLOAD auditor ke folder penugasan.
Output ditulis ke salinan kerja `_KKP/LKE-terisi-<skill>.xlsx` (template/upload
asli tidak pernah diubah).
"""
import json
import re
import shutil
from pathlib import Path

from claude_agent_sdk import tool
from openpyxl import load_workbook

from app.config import get_settings
from app.lke_writer import LKEWriter

# Sheet agregator SPIP (formula-only lead) — tak boleh ditulis. Nama mengikuti
# template rev4 2025 (KKlead I KL → KKLEAD I).
_SPIP_AGGREGATORS = {"KKLEAD_SPIP", "KKLEAD I", "KKLEAD II", "KKLEAD III"}


def _slug(skill: str) -> str:
    return re.sub(r"[^a-z0-9\-]", "-", str(skill).strip().lower())


def _spip_template() -> tuple[Path, Path]:
    base = get_settings().skills_path / "evaluasi-spip" / "references" / "templates"
    return base / "lke-spip-kementerian.xlsx", base / "cell-map-formulas.json"


def _find_uploaded_xlsx(folder: Path) -> Path | None:
    """Cari LKE .xlsx yang diupload auditor di subfolder input penugasan.

    Abaikan output kerja (_KKP/_LHP/_QA-SAIPI) supaya tidak memungut hasil sendiri.
    """
    skip = {"_KKP", "_LHP", "_QA-SAIPI", "_INGESTED"}
    candidates: list[Path] = []
    for p in folder.rglob("*.xlsx"):
        if any(part in skip for part in p.relative_to(folder).parts):
            continue
        if p.name.startswith("~$"):  # lock file Excel
            continue
        candidates.append(p)
    return sorted(candidates)[0] if candidates else None


def _resolve_source(folder: Path, skill: str) -> tuple[Path | None, Path | None, str]:
    """Tentukan (source_xlsx, cellmap, note). Prefer upload auditor; SPIP fallback ke template."""
    uploaded = _find_uploaded_xlsx(folder)
    if uploaded is not None:
        return uploaded, None, f"LKE dari upload: {uploaded.name}"
    if _slug(skill) == "evaluasi-spip":
        tpl, cmap = _spip_template()
        if tpl.is_file():
            return tpl, (cmap if cmap.is_file() else None), f"LKE template SPIP: {tpl.name}"
    return None, None, "tidak ada LKE — upload file .xlsx LKE dulu"


@tool(
    "fill_lke",
    "Isi Lembar Kerja Evaluasi (LKE) Excel untuk skill evaluasi (SAKIP/SPIP) TANPA "
    "mengubah rumus — hanya cell INPUT yang ditulis; cell formula & sheet agregator "
    "DITOLAK otomatis (dilaporkan di 'refused'). Sumber: LKE .xlsx yang diupload "
    "auditor, atau template SPIP bawaan. Output: _KKP/LKE-terisi-<skill>.xlsx (asli "
    "tak diubah). `entries` = list of {sheet, coord (mis. 'K6'), value, note?}. "
    "Pakai SEBELUM menyusun catatan/temuan untuk skill ber-LKE.",
    {"penugasan_folder": str, "skill": str, "entries": list},
)
async def fill_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = str(args.get("skill", "")).strip()
    entries = args.get("entries") or []
    if not isinstance(entries, list) or not entries:
        return {"content": [{"type": "text", "text": "FAILED|entries kosong (list of {sheet,coord,value})"}], "is_error": True}

    src, cellmap, note = _resolve_source(folder, skill)
    if src is None:
        return {"content": [{"type": "text", "text": f"FAILED|{note}"}], "is_error": True}

    out = folder / "_KKP" / f"LKE-terisi-{_slug(skill)}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    # Mulai dari salinan kerja yang sudah ada (akumulatif) bila ada, else dari source.
    base = out if out.is_file() else src
    try:
        writer = LKEWriter(
            base, cellmap_path=cellmap,
            aggregator_sheets=_SPIP_AGGREGATORS if _slug(skill) == "evaluasi-spip" else None,
        )
    except Exception as e:  # noqa: BLE001
        return {"content": [{"type": "text", "text": f"FAILED|gagal buka LKE: {e}"}], "is_error": True}

    for e in entries:
        if not isinstance(e, dict):
            continue
        writer.set(str(e.get("sheet", "")), str(e.get("coord", "")), e.get("value"), e.get("note"))
    writer.save(out)

    payload = {
        "ok": True,
        "sumber": note,
        "output": str(out.relative_to(folder)),
        "ditulis": len(writer.written),
        "ditolak_count": len(writer.refused),
        "ditolak_formula": writer.refused[:40],  # cell ditolak (formula/agregator) — pilih cell input lain
    }
    return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}


_EVAL_LABELS_HDR = ("EVALUASI", "EVALUASI APIP")


def _baca_fokus(ws, mulai: int = 1, cap_baris: int = 120) -> dict:
    """Baca HANYA baris data hidup × kolom relevan (uraian + blok PM + blok PK).

    Sheet LKE besar (mis. KK 5.2) memuat puluhan ribu sel formula/format —
    membaca SEMUA sel tak praktis pada cap berapa pun. Untuk menilai, agen
    sebenarnya hanya perlu: uraian kriteria (kolom B), nilai PM auditee, dan
    kolom PK yang harus diisi. Satu kali lintas, ramah mode read-only.
    """
    maks = ws.max_row or 0
    # Sheet LKE bisa SANGAT lebar — blok PK KK3.x ada di kolom 103-110, jadi
    # batas kolom harus longgar (80 kolom membuat blok PK tak terdeteksi).
    lebar = min(ws.max_column or 120, 220)
    pm_col = pk_col = ev_col = None
    header_row = 0
    counter_row = 0
    started = False
    baris: list[dict] = []
    lanjut: int | None = None

    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=maks, min_col=1, max_col=lebar), start=1
    ):
        if pk_col is None and idx <= 8:
            for c in row:
                v = str(c.value or "").strip().upper()
                if v == "PM" and pm_col is None:
                    pm_col, header_row = c.column, idx
                elif v == "PK" and pk_col is None:
                    pk_col, header_row = c.column, (header_row or idx)
                elif v in _EVAL_LABELS_HDR and ev_col is None and pk_col and c.column > pk_col:
                    ev_col = c.column
            continue
        if pk_col is None:
            continue

        vals = [c.value for c in row]
        # Baris "counter" auto-nomor (=X{n}+1) memisahkan header dari data.
        if not started and any(
            isinstance(v, str) and v.startswith("=") and "+1" in v for v in vals[:3]
        ):
            counter_row = idx
            continue
        if idx <= counter_row:
            continue

        b = vals[1] if len(vals) > 1 else None
        if b in (None, "") or (isinstance(b, str) and b.startswith("=")):
            if started:
                break          # blok data kontigu selesai
            continue
        bs = str(b).strip().lower()
        if bs.startswith("petunjuk") or re.match(r"kolom\s+\d+\s*:", bs):
            break              # masuk area legenda/petunjuk
        started = True
        if idx < mulai:
            continue
        if len(baris) >= cap_baris:
            lanjut = idx
            break

        pm_end = (pk_col - 1) if (pm_col and pk_col > pm_col) else None
        pk_end = (ev_col - 1) if ev_col else min(pk_col + 4, lebar)

        def _blok(s, e):
            out = {}
            if not s:
                return out
            for cc in row[s - 1:e]:
                if cc.value not in (None, ""):
                    out[cc.coordinate] = str(cc.value)[:60]
            return out

        pk_cells = _blok(pk_col, pk_end)
        baris.append({
            "row": idx,
            "uraian": str(b)[:200],
            "pm": _blok(pm_col, pm_end),
            "pk": pk_cells,
            "pk_terisi": bool(pk_cells),
        })

    return {
        "mode": "fokus",
        "header_row": header_row,
        "kolom_pm": pm_col,
        "kolom_pk": [pk_col, (ev_col - 1) if ev_col else None] if pk_col else None,
        "n_baris": len(baris),
        "baris": baris,
        "ada_lanjutan": lanjut is not None,
        "lanjut_dari_baris": lanjut,
        "catatan": ("Hanya baris data hidup × kolom relevan (uraian + PM + PK). "
                    "Isi kolom PK via fill_lke pada koordinat di blok 'pk'."),
    }


def _lke_to_read(folder: Path, skill: str) -> tuple[Path | None, str]:
    """LKE yang dibaca: salinan kerja _KKP/LKE-terisi (bila ada, paling mutakhir),
    else sumber (upload auditee / template)."""
    work = folder / "_KKP" / f"LKE-terisi-{_slug(skill)}.xlsx"
    if work.is_file():
        return work, f"salinan kerja: {work.name}"
    src, _cmap, note = _resolve_source(folder, skill)
    return src, note


@tool(
    "read_lke",
    "Baca isi Lembar Kerja Evaluasi (LKE) auditee untuk skill evaluasi (SAKIP/SPIP) — "
    "self-assessment yang sudah diisi auditee. Tanpa `sheet`: daftar nama sheet + "
    "jumlah cell terisi. Dengan `sheet`: nilai cell non-kosong (coord→{v,f}; f=true bila "
    "FORMULA, jangan ditulis). Pakai untuk MENILAI self-assessment auditee sebelum "
    "mengisi kolom APIP via fill_lke. "
    "**`fokus=1` (DISARANKAN untuk menilai)** — kembalikan hanya BARIS DATA HIDUP × "
    "kolom relevan: `uraian` (kriteria), blok `pm` (nilai auditee), blok `pk` (kolom "
    "APIP yang harus diisi) + `pk_terisi`. Ini satu-satunya cara praktis pada sheet "
    "raksasa (mis. KK 5.2: puluhan ribu sel formula — mode biasa tak akan tuntas). "
    "Koordinat di blok `pk` langsung dipakai untuk `fill_lke`. "
    "**Mode biasa** (tanpa fokus) membaca semua sel per JENDELA BARIS (≤600 sel/"
    "panggilan): `mulai_baris` + `jumlah_baris`. Bila `ada_lanjutan=true`, panggil LAGI "
    "dengan `mulai_baris=lanjut_dari_baris` sampai `ada_lanjutan=false` — JANGAN "
    "menyimpulkan penilaian sebelum seluruh baris terbaca.",
    {"penugasan_folder": str, "skill": str, "sheet": str,
     "mulai_baris": int, "jumlah_baris": int, "fokus": int},
)
async def read_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = str(args.get("skill", "")).strip()
    sheet = str(args.get("sheet", "")).strip()
    src, note = _lke_to_read(folder, skill)
    if src is None or not Path(src).is_file():
        return {"content": [{"type": "text", "text": f"FAILED|{note}"}], "is_error": True}
    try:
        # read_only=True: LKE SPIP rev4 (7 MB, 28 sheet) butuh ~11 detik dgn mode
        # normal vs ~0,1 detik read-only. Tool ini hanya MEMBACA, dan penelusuran
        # per-jendela berarti beberapa panggilan → mode normal bikin run melar.
        wb = load_workbook(src, data_only=False, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {"content": [{"type": "text", "text": f"FAILED|gagal buka LKE: {e}"}], "is_error": True}

    if not sheet:
        sheets = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            filled = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
            sheets.append({"sheet": sn, "terisi": filled, "dim": ws.dimensions})
        payload = {"sumber": note, "total_sheet": len(sheets), "sheets": sheets}
        return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}

    if sheet not in wb.sheetnames:
        return {"content": [{"type": "text", "text": f"NOT_FOUND|sheet '{sheet}'. Ada: {wb.sheetnames}"}], "is_error": True}
    ws = wb[sheet]

    # MODE FOKUS — untuk sheet raksasa (puluhan ribu sel formula/format), membaca
    # semua sel tak praktis. Kembalikan hanya baris data hidup × kolom relevan.
    if str(args.get("fokus") or "").strip() not in ("", "0", "false", "False"):
        hasil = _baca_fokus(ws, mulai=max(1, int(args.get("mulai_baris") or 1)))
        hasil["sumber"] = note
        hasil["sheet"] = sheet
        return {"content": [{"type": "text", "text": json.dumps(hasil, ensure_ascii=False)}]}

    # Dibaca per JENDELA BARIS supaya sheet raksasa (LKE SPIP rev4: ribuan sel)
    # bisa ditelusuri TUNTAS lewat beberapa panggilan. Dulu mentok 150 sel PERTAMA
    # lalu berhenti — sisa sheet tak pernah terlihat agen (baris PM tak ternilai).
    _CAP = 600  # batas sel per panggilan (menjaga ukuran payload tetap wajar)
    mulai = max(1, int(args.get("mulai_baris") or 1))
    maks_baris = ws.max_row or 0
    # Jendela dibatasi JUMLAH SEL, bukan jumlah baris: sheet LKE sering melaporkan
    # max_row raksasa (mis. 15.008) padahal baris berisi jauh lebih sedikit —
    # membatasi per-baris akan memboroskan panggilan pada area kosong.
    jb = int(args.get("jumlah_baris") or 0)
    akhir = (mulai + max(1, min(jb, 5000)) - 1) if jb > 0 else maks_baris

    cells: dict[str, dict] = {}
    capped = False
    baris_terakhir = 0
    if mulai <= maks_baris:
        for row in ws.iter_rows(min_row=mulai, max_row=min(akhir, maks_baris)):
            for c in row:
                if c.value in (None, ""):
                    continue
                if len(cells) >= _CAP:
                    capped = True
                    break
                cells[c.coordinate] = {"v": str(c.value)[:60], "f": c.data_type == "f"}
                baris_terakhir = max(baris_terakhir, c.row)
            if capped:
                break

    # Bila terpotong cap → sambung dari baris terakhir yang terbaca (bukan lompat).
    lanjut = (baris_terakhir + 1) if capped else (akhir + 1)
    # `max_row` sering meleset JAUH (mis. KK 5.2 melaporkan 15.008 baris padahal
    # data nyata berhenti di baris awal — sisanya artefak format/validasi). Kalau
    # `ada_lanjutan` dipatok ke max_row, agen akan menyusuri belasan ribu baris
    # kosong. Jadi: cek apakah MASIH ADA DATA di depan, bukan sekadar sisa baris.
    # Look-ahead DIBATASI (_LOOKAHEAD baris) supaya murah: cukup untuk memastikan
    # tidak berhenti di tengah data, tanpa memindai belasan ribu baris kosong.
    _LOOKAHEAD = 500
    ada_lanjutan = False
    if lanjut <= maks_baris:
        batas = min(maks_baris, lanjut + _LOOKAHEAD - 1)
        for row in ws.iter_rows(min_row=lanjut, max_row=batas):
            if any(c.value not in (None, "") for c in row):
                ada_lanjutan = True
                break
    payload = {
        "sumber": note,
        "sheet": sheet,
        "baris_mulai": mulai,
        "baris_akhir_dibaca": baris_terakhir or min(akhir, maks_baris),
        "baris_maks_sheet": maks_baris,
        "n_cell": len(cells),
        "capped": capped,
        "ada_lanjutan": ada_lanjutan,
        "lanjut_dari_baris": lanjut if ada_lanjutan else None,
        "cells": cells,
    }
    if ada_lanjutan:
        payload["petunjuk"] = (
            f"Sheet belum habis — panggil read_lke lagi dengan mulai_baris={lanjut} "
            f"(sampai ada_lanjutan=false) sebelum menyimpulkan penilaian."
        )
    return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}


@tool(
    "write_penilaian_lke",
    "Tulis (overwrite) _KKP/penilaian-lke-<skill>.json — REKAP skor/predikat hasil "
    "evaluasi ber-LKE (SAKIP/SPIP), sumber tunggal untuk rekap di KKP. Panggil SEKALI "
    "setelah SEMUA unsur/komponen dinilai & fill_lke selesai. Struktur `penilaian` = "
    "{komponen:[{nama, bobot, nilai_pm, nilai_apip, predikat, catatan?}], total_pm?, "
    "total_apip?, predikat_akhir?}. nilai_pm = penilaian mandiri auditee, nilai_apip = "
    "hasil penjaminan APIP. render_kkp_docx akan menampilkan tabel 'Rekap Penilaian (LKE)' "
    "dari file ini di atas daftar AoI.",
    {"penugasan_folder": str, "skill": str, "penilaian": dict},
)
async def write_penilaian_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = _slug(args.get("skill", ""))
    pen = args.get("penilaian") or {}
    if not skill:
        return {"content": [{"type": "text", "text": "FAILED|skill kosong"}], "is_error": True}
    out_dir = folder / "_KKP"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"penilaian-lke-{skill}.json"
    pen.setdefault("skill", skill)
    out.write_text(json.dumps(pen, ensure_ascii=False, indent=2), encoding="utf-8")
    n = len(pen.get("komponen", []) or [])
    return {"content": [{"type": "text", "text": f"OK|penilaian-lke ditulis|n_komponen={n}|{out.relative_to(folder)}"}]}


# =============================================================================
# BATCH EVALUASI SPIP — LKE SPIP volumenya besar (25 sub-unsur, ~24 sheet, ribuan
# sel APIP). Satu run agen sering kehabisan budget sebelum tuntas → sheet volume
# tinggi ditinggalkan kosong. Solusi: kerjakan per BATCH (per komponen/KK Lead),
# satu batch per run, dengan gate kelengkapan deterministik.
# =============================================================================

# Sheet detail (diisi blok PK) dikelompokkan ke 3 batch = KK Lead I/II/III.
# Nama sheet mengikuti template rev4 2025 (multi-satker). Sheet lead/agregator
# (KKLEAD*) terhitung otomatis dari rumus → tidak masuk batch.
SPIP_BATCHES = [
    {"no": 1, "nama": "KK Lead I — Penetapan Tujuan",
     "sheets": ["KKE 1 SASTRA", "KKE 2.1 SASPRO", "KKE 2.2 SASKEG", "KKE 2.3 SAS RO"]},
    {"no": 2, "nama": "KK Lead II — Struktur & Proses",
     "sheets": ["KK3.1", "KK3.2", "KK3.3", "KK3.4"]},
    {"no": 3, "nama": "KK Lead III — Pencapaian Tujuan",
     "sheets": ["KK 5.1 A", "KK 5.1 B", "KK 5.2", "KK 6", "KK 7", "KK 8", "KK 4"]},
]

# Ambang: sheet dianggap lengkap bila >=95% baris hidup (kolom B terisi) sudah ada
# isian di blok PK. Tinggi karena tujuannya ANTI-KEBOLONGAN (semua baris PM harus
# dinilai). Kolom label blok penilaian di baris atas: PM | PK | EVALUASI.
_BATCH_THRESHOLD = 0.95
_PK_LABELS = ("PK",)
_EVAL_LABELS = ("EVALUASI", "EVALUASI APIP")


def _find_pk_block(ws) -> tuple[int, int, int] | None:
    """Cari blok kolom PK (penilaian APIP) → (pk_start, pk_end, header_row).

    Template rev4: baris atas memuat label blok `PM | PK | EVALUASI`. Blok PK =
    kolom dari label 'PK' sampai sebelum 'EVALUASI'. Return None bila tak ada
    label PK (sheet tak terukur → panduan-saja, tak di-gate keras).
    """
    pk_col = eval_col = None
    header_row = 0
    for r in range(1, 7):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().upper()
            if v in _PK_LABELS and pk_col is None:
                pk_col = c
                header_row = r
            elif v in _EVAL_LABELS and eval_col is None and pk_col is not None and c > pk_col:
                eval_col = c
    if pk_col is None:
        return None
    pk_end = (eval_col - 1) if eval_col else min(pk_col + 4, ws.max_column)
    return (pk_col, pk_end, header_row)


def _sheet_apip_ratio(ws) -> tuple[float, int, int]:
    """(ratio, filled_rows, total_data_rows) kelengkapan blok PK satu sheet.

    Baris data = baris di bawah header yang **kolom B**-nya berisi & bukan rumus
    (aturan auditor: baris hidup ditandai isi kolom B dari PM satker). Baris
    dianggap terisi PK bila minimal satu sel di blok PK non-kosong.
    ratio = -1 bila (a) tak ada blok PK, atau (b) tak ada baris hidup (kolom B).
    """
    blk = _find_pk_block(ws)
    if blk is None:
        return (-1.0, 0, 0)
    pk_start, pk_end, hr = blk
    # Baris DATA mulai setelah baris "counter" auto-nomor (salah satu kolom A–C
    # berupa rumus `=X{n}+1`). Ini memisahkan baris HEADER (yang kolom B-nya berisi
    # label, mis. "URAIAN SASARAN STRATEGIS") dari baris data asli → hindari
    # false positive header terhitung sebagai baris hidup.
    data_start = hr + 1
    for r in range(hr + 1, min(hr + 12, ws.max_row + 1)):
        if any(isinstance(ws.cell(r, c).value, str)
               and ws.cell(r, c).value.startswith("=") and "+1" in ws.cell(r, c).value
               for c in (1, 2, 3)):
            data_start = r + 1
            break
    # Hitung hanya BLOK DATA KONTIGU dari data_start. Berhenti di baris kosong
    # (akhir blok) atau marker legend "Petunjuk"/"Kolom N :" — blok instruksi di
    # bawah sheet (kolom B berisi teks) JANGAN ikut terhitung sebagai baris data.
    total = filled = 0
    started = False
    for r in range(data_start, ws.max_row + 1):
        b = ws.cell(r, 2).value
        b_empty = b in (None, "") or (isinstance(b, str) and b.startswith("="))
        if b_empty:
            if started:
                break
            continue
        bs = str(b).strip().lower()
        if bs.startswith("petunjuk") or re.match(r"kolom\s+\d+\s*:", bs):
            break
        started = True
        total += 1
        if any(ws.cell(r, c).value not in (None, "") for c in range(pk_start, pk_end + 1)):
            filled += 1
    if total == 0:
        return (-1.0, 0, 0)
    return (filled / total, filled, total)


# ── Dukungan struktur LKE LAMA (pra-rev4) ─────────────────────────────────────
# Menambah cabang tanpa mengubah jalur rev4. Batch struktur lama pakai nama sheet
# lama (KKE 1.1 SASTRA PEMDA, KKlead I KL, dst; ada trailing space di KK 5.x).
SPIP_BATCHES_OLD = [
    {"no": 1, "nama": "KK Lead I — Penetapan Tujuan",
     "sheets": ["KKE 1.1 SASTRA PEMDA", "KKE 1.2 SASARAN OPD", "KKE 2.1 SASKEG",
                "KK 2.2 RO", "KKE 2.2 KEGIATAN"]},
    {"no": 2, "nama": "KK Lead II — Struktur & Proses",
     "sheets": ["KK3.1", "KK3.2", "KK3.3", "KK3.4"]},
    {"no": 3, "nama": "KK Lead III — Pencapaian Tujuan",
     "sheets": ["KK 5.1A", "KK 5.1 B ", "KK 5.2 ", "KK 6", "KK 7", "KK 8", "KK4_PENALTI"]},
]

_APIP_HEADER_HINTS_OLD = ("PENJAMIN", "EVALUASI APIP", "NILAI PK", "PENJAMINAN")


def _sheet_apip_ratio_old(ws) -> tuple[float, int, int]:
    """Detektor struktur LAMA: kolom penilaian dari header 'PENJAMIN/NILAI PK'.
    Baris data = kolom A/B/C berisi. Kasar tapi cukup untuk MEMICU batching
    (mencegah overload) pada LKE lama. ratio=-1 bila kolom APIP tak terdeteksi."""
    apip_cols: list[int] = []
    header_row = 0
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(h in v.upper() for h in _APIP_HEADER_HINTS_OLD):
                apip_cols.append(c)
                header_row = max(header_row, r)
    apip_cols = sorted(set(apip_cols))
    if not apip_cols:
        return (-1.0, 0, 0)
    total = filled = 0
    for r in range(header_row + 1, ws.max_row + 1):
        if not any(ws.cell(r, c).value not in (None, "") for c in (1, 2, 3)):
            continue
        total += 1
        if any(ws.cell(r, c).value not in (None, "") for c in apip_cols):
            filled += 1
    if total == 0:
        return (-1.0, 0, 0)
    return (filled / total, filled, total)


def _detect_spip_structure(wb) -> str:
    """'rev4' (KKE 1 SASTRA / KKLEAD I) atau 'old' (KKE 1.1 SASTRA PEMDA /
    KKlead I KL). Default 'rev4' bila tak jelas."""
    names = set(wb.sheetnames)
    if "KKE 1 SASTRA" in names or "KKLEAD I" in names:
        return "rev4"
    if "KKE 1.1 SASTRA PEMDA" in names or "KKlead I KL" in names:
        return "old"
    return "rev4"


def spip_batch_progress(folder: Path) -> dict:
    """Status kelengkapan per batch SPIP dari `_KKP/LKE-terisi-evaluasi-spip.xlsx`.

    Deterministik (no-LLM). Dipakai tool `lke_batch_status` + gate render_kkp_docx.
    Tiap sheet ter-gate melapor terisi/total/sisa (sisa = baris hidup yang blok
    PK-nya masih kosong) → mendukung batch sambung lintas run untuk sheet raksasa.
    """
    src = folder / "_KKP" / "LKE-terisi-evaluasi-spip.xlsx"
    if not src.is_file():
        return {
            "exists": False, "all_complete": False, "next_batch": SPIP_BATCHES[0]["no"],
            "batches": [{"no": b["no"], "nama": b["nama"], "complete": False,
                         "sisa_baris": None, "sheets": {}} for b in SPIP_BATCHES],
        }
    try:
        wb = load_workbook(src, data_only=False, read_only=False)
    except Exception as e:  # noqa: BLE001
        return {"exists": True, "error": str(e)[:200], "all_complete": False, "next_batch": 1, "batches": []}

    # Deteksi struktur → pilih definisi batch + detektor yang cocok (rev4 tetap
    # jalur utama; struktur lama ditambahkan agar LKE lama tetap bisa dibatch).
    struktur = _detect_spip_structure(wb)
    batches_def = SPIP_BATCHES if struktur == "rev4" else SPIP_BATCHES_OLD
    ratio_fn = _sheet_apip_ratio if struktur == "rev4" else _sheet_apip_ratio_old

    batches_out: list[dict] = []
    next_batch: int | None = None
    any_measured = False  # minimal 1 sheet punya baris hidup (kolom B) + blok PK
    for b in batches_def:
        sheet_stat: dict[str, dict] = {}
        incomplete = False   # ada sheet terukur yang belum lengkap
        sisa_batch = 0       # total baris hidup yang belum ada PK (untuk sambung)
        for sn in b["sheets"]:
            if sn not in wb.sheetnames:
                continue
            ratio, filled, total = ratio_fn(wb[sn])
            if ratio < 0:
                sheet_stat[sn] = {"status": "panduan-saja"}  # tak ada blok PK / baris hidup
                continue
            any_measured = True
            ok = ratio >= _BATCH_THRESHOLD
            sisa = total - filled
            sheet_stat[sn] = {"terisi": filled, "total": total, "sisa": sisa,
                              "pct": round(ratio * 100), "ok": ok}
            if not ok:
                incomplete = True
                sisa_batch += sisa
        complete = not incomplete
        batches_out.append({"no": b["no"], "nama": b["nama"], "complete": complete,
                            "sisa_baris": sisa_batch, "sheets": sheet_stat})
        if incomplete and next_batch is None:
            next_batch = b["no"]
    # Guard anti false-pass: bila TIDAK ADA baris hidup terdeteksi (mis. LKE masih
    # kosong / PM belum diisi satker), JANGAN nyatakan lengkap.
    if not any_measured:
        return {"exists": True, "struktur": struktur, "batches": batches_out,
                "next_batch": batches_def[0]["no"], "all_complete": False, "any_measured": False,
                "catatan": "Tidak ada baris hidup terdeteksi — pastikan LKE berisi PM "
                           "satker sebelum mengisi PK."}
    return {"exists": True, "struktur": struktur, "batches": batches_out, "next_batch": next_batch,
            "all_complete": next_batch is None, "any_measured": True}


@tool(
    "lke_batch_status",
    "Cek progres BATCH evaluasi SPIP (khusus evaluasi-spip). LKE SPIP volumenya besar "
    "→ dikerjakan per BATCH (per komponen/KK Lead), satu batch per run. Panggil DI AWAL "
    "tiap run untuk tahu batch mana yang harus dikerjakan (next_batch + daftar sheet-nya), "
    "dan SETELAH fill_lke untuk memastikan batch itu lengkap (semua sheet ok) sebelum "
    "berhenti & memberi reminder ke auditor. Return JSON: batches[] (per sheet: %terisi "
    "kolom APIP + ok), next_batch, all_complete. Skill selain evaluasi-spip → applicable=false.",
    {"penugasan_folder": str, "skill": str},
)
async def lke_batch_status(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = _slug(args.get("skill", ""))
    if skill != "evaluasi-spip":
        return {"content": [{"type": "text", "text": json.dumps(
            {"applicable": False, "skill": skill,
             "catatan": "Batch hanya untuk evaluasi-spip. SAKIP/lainnya 1-lintasan."})}]}
    prog = spip_batch_progress(folder)
    prog["applicable"] = True
    return {"content": [{"type": "text", "text": json.dumps(prog, ensure_ascii=False)}]}


LKE_TOOLS = [read_lke, fill_lke, write_penilaian_lke, lke_batch_status]
